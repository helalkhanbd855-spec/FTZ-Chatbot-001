from flask import Flask, request, jsonify
from flask_cors import CORS
import datetime
import os
import json
from openpyxl import Workbook, load_workbook
import openai

app = Flask(__name__)
CORS(app)

openai.api_key = os.getenv("OPENAI_API_KEY")

# In-memory chat state
user_states = {}

# ---------- Helper Functions ----------

def get_time_greeting():
    hour = datetime.datetime.now().hour
    if 5 <= hour < 12:
        return "Good morning"
    elif 12 <= hour < 17:
        return "Good afternoon"
    elif 17 <= hour < 21:
        return "Good evening"
    else:
        return "Good night"

def load_users():
    """Load or create users.json"""
    if not os.path.exists("users.json"):
        with open("users.json", "w") as f:
            json.dump({}, f)
    with open("users.json", "r") as f:
        return json.load(f)

def save_users(users):
    with open("users.json", "w") as f:
        json.dump(users, f, indent=2)

def save_to_excel(filename, sheet_name, headers, row_data):
    """Create or append a record into a user-specific Excel sheet."""
    file_exists = os.path.exists(filename)

    if file_exists:
        wb = load_workbook(filename)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            ws.append(headers)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(headers)

    ws.append(row_data)
    wb.save(filename)

# ---------- LOGIN / REGISTER ----------

@app.route("/register", methods=["POST"])
def register():
    data = request.get_json()
    username = data.get("username")
    password = data.get("password")

    users = load_users()

    if username in users:
        return jsonify({"status": "error", "message": "Username already exists!"})

    users[username] = {"password": password}
    save_users(users)
    return jsonify({"status": "success", "message": "Registration successful!"})


@app.route("/login", methods=["POST"])
def login():
    data = request.get_json()
    username = data.get("username")
    password = data.get("password")

    users = load_users()

    if username not in users or users[username]["password"] != password:
        return jsonify({"status": "error", "message": "Invalid credentials!"})

    return jsonify({"status": "success", "message": f"Welcome {username}!"})

# ---------- CHATBOT ----------

@app.route('/chat', methods=['POST'])
def chat():
    data = request.get_json() or {}
    msg = data.get('message', '').strip().lower()
    username = data.get('username', 'Guest')

    greeting = get_time_greeting()
    reply = ""

    state = user_states.get(username, {"mode": None, "step": None, "data": {}})

    # ------------------- SITE UPDATE FLOW -------------------
    if state["mode"] == "site_update":
        if state["step"] == "ask_site":
            state["data"]["site_name"] = msg
            reply = "What job done today?"
            state["step"] = "ask_job"

        elif state["step"] == "ask_job":
            state["data"]["job_done"] = msg
            reply = "What balance remains?"
            state["step"] = "ask_balance"

        elif state["step"] == "ask_balance":
            state["data"]["balance"] = msg
            reply = "How many persons worked?"
            state["step"] = "ask_person"

        elif state["step"] == "ask_person":
            state["data"]["person_count"] = msg
            reply = "Any other update?"
            state["step"] = "ask_other"

        elif state["step"] == "ask_other":
            state["data"]["other_update"] = msg

            now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            save_to_excel(
                "Site_Updates.xlsx",
                username,
                ["DateTime", "User", "Site", "Job Done", "Balance", "Persons", "Other Update"],
                [
                    now,
                    username,
                    state["data"].get("site_name", ""),
                    state["data"].get("job_done", ""),
                    state["data"].get("balance", ""),
                    state["data"].get("person_count", ""),
                    state["data"].get("other_update", "")
                ]
            )

            reply = (
                f"✅ {username}, your site update has been submitted!\n\n"
                f"📅 Date: {now}\n"
                f"📍 Site: {state['data']['site_name']}\n"
                f"🧰 Job: {state['data']['job_done']}\n"
                f"📊 Balance: {state['data']['balance']}\n"
                f"👷 Persons: {state['data']['person_count']}\n"
                f"🗒️ Other: {state['data']['other_update']}\n\n"
                f"Saved in sheet '{username}' ✅"
            )

            user_states[username] = {"mode": None, "step": None, "data": {}}
            return jsonify({'reply': reply})

        user_states[username] = state
        return jsonify({'reply': reply})

    # ------------------- TIMECARD UPDATE FLOW -------------------
    if state["mode"] == "timecard_update":
        if state["step"] == "ask_time":
            work_hours = msg
            now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            save_to_excel(
                "Timecards.xlsx",
                username,
                ["DateTime", "User", "Work Hours"],
                [now, username, work_hours]
            )

            reply = f"✅ {username}, your timecard is submitted! ({work_hours})"
            user_states[username] = {"mode": None, "step": None, "data": {}}
            return jsonify({'reply': reply})

    # ------------------- MAIN CHAT LOGIC -------------------
    if not msg:
        reply = "Please say something 🙂"

    elif any(word in msg for word in ["hi", "hello", "hey"]):
        reply = (
            f"{greeting}, {username}! 👋\n"
            "How can I help you today?\n"
            "Type 'know' to learn about systems or 'update' to give new info."
        )

    elif "update" in msg:
        if "time" in msg:
            reply = "Please enter your work hours (e.g. '9 AM to 6 PM')."
            user_states[username] = {"mode": "timecard_update", "step": "ask_time", "data": {}}
        elif "site" in msg:
            reply = "Let's start your site update. Which site?"
            user_states[username] = {"mode": "site_update", "step": "ask_site", "data": {}}
        else:
            reply = "Please choose: 1️⃣ Timecard update or 2️⃣ Site update."

    elif "bye" in msg:
        reply = f"Goodbye {username}! 👋 Have a great day."

    # ---------- OpenAI GPT fallback ----------
    elif not reply:
        try:
            completion = openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=[{"role": "user", "content": msg}]
            )
            reply = completion.choices[0].message.content
        except Exception as e:
            reply = f"⚠️ ChatGPT Error: {e}"

    return jsonify({'reply': reply})


# ---------- RUN APP ----------
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
